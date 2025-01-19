# %%
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

"""
STEP 1: Read and Process Input Data
"""
# Specify the path to your input Excel file
# Students: Replace this path with your actual file path
input_file_path = 'Input Data/PO_Report.xlsx'  # Use relative path for better portability

# Read the Excel file into a pandas DataFrame
try:
    df_po_report = pd.read_excel(input_file_path)
    print("Successfully loaded the input file.")
    print("\nFirst few rows of the data:")
    print(df_po_report.head())
except Exception as e:
    print(f"Error loading the file: {str(e)}")
    exit()

"""
STEP 2: Data Preprocessing
"""
# Convert 'Po Creation Date' to datetime format for proper sorting
df_po_report['Po Creation Date'] = pd.to_datetime(df_po_report['Po Creation Date'])

# Sort the dataframe by PO Creation Date (ascending order)
df_po_report = df_po_report.sort_values(by='Po Creation Date', ascending=True)

# Remove rows where PO Qty Due is 0 (filtering out completed POs)
df_po_report = df_po_report[df_po_report['PO Qty Due'] != 0]

# Get list of unique supplier names for processing
unique_suppliers = df_po_report['Supplier Name'].unique()
print(f"\nNumber of unique suppliers found: {len(unique_suppliers)}")

"""
STEP 3: Set up Output Directory Structure
"""
# Define the output folder structure
output_folder = os.path.join('Output Data', 'Supplier_PO_Reports')

# Create output folders if they don't exist
if not os.path.exists(output_folder):
    os.makedirs(output_folder)
    print(f"\nCreated output directory: {output_folder}")

# Create a timestamped subfolder for this run
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
dated_folder = os.path.join(output_folder, f'PO_Reports_{timestamp}')
os.makedirs(dated_folder)

def format_excel_file(filename):
    """
    Function to apply professional formatting to the Excel file
    
    Parameters:
    filename (str): Path to the Excel file to be formatted
    """
    wb = load_workbook(filename)
    ws = wb.active
    
    # Define styling elements
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')  # Dark blue header
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')  # White bold text for header
    regular_font = Font(name='Calibri', size=10)  # Regular text for data
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for cell in ws[1]:  # First row
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Format data rows
    data_rows = list(ws.rows)[1:]  # All rows except header
    for row in data_rows:
        for cell in row:
            cell.font = regular_font
            cell.border = thin_border
            
            # Right-align numbers, left-align text
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # Find the maximum length of content in each column
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set column width (with a maximum of 50)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Save the formatted workbook
    wb.save(filename)

"""
STEP 5: Generate Individual Supplier Reports
"""
# Process each supplier
for supplier in unique_suppliers:
    # Filter data for current supplier
    supplier_df = df_po_report[df_po_report['Supplier Name'] == supplier]
    
    # Create a safe filename (remove special characters)
    safe_supplier_name = "".join(x for x in supplier if x.isalnum() or x in (' ', '-', '_'))
    
    # Generate filename with timestamp
    excel_filename = f"{safe_supplier_name}_PO_Report_{timestamp}.xlsx"
    
    # Create full output path
    output_path = os.path.join(dated_folder, excel_filename)
    
    try:
        # Save and format the supplier's report
        supplier_df.to_excel(output_path, index=False)
        format_excel_file(output_path)
        print(f"Created and formatted PO report for {supplier}: {excel_filename}")
    except Exception as e:
        print(f"Error creating report for {supplier}: {str(e)}")

print(f"\nProcess completed! All formatted supplier reports have been generated in: {dated_folder}")

"""
Additional Notes for Students:
1. Make sure you have all required libraries installed:
   pip install pandas openpyxl

2. Expected Input Excel Format:
   - Must have columns: 'Po Creation Date', 'PO Qty Due', 'Supplier Name'
   - Dates should be in a recognizable format

3. Output:
   - Creates a timestamped folder for each run
   - Generates individual Excel files for each supplier
   - Applies professional formatting to each report

4. Customization Options:
   - Change header_fill color by modifying the color code
   - Adjust font sizes in the format_excel_file function
   - Modify column width limits (currently max 50)
"""


